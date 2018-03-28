
import openpyxl
import os
import xlsxwriter

# Classes #


class Volunteer:
    def __init__(self, name, email, phone, contactMethod, allergies):
        self.name           = name
        self.email          = email
        self.phone          = phone
        self.contactMethod  = contactMethod
        self.allergies      = allergies


class Group:
    def __init__(self, members, day1, day2):
        self.members    = members
        self.day1       = day1
        self.day2       = day2
        self.allergies  = []
        # automatically collects allergies from the members and assigns to overall group
        for member in self.members:
            for allergy in member.allergies:
                if allergy not in self.allergies:
                    self.allergies.append(allergy)

    def __len__(self):
        return len(self.members)
    
    def __add__(self, otherGroup):
        newGroup = Group(self.members + otherGroup.members, self.day1, 5)
        return newGroup

    def has_member(self, volunteer):
        return volunteer in self.members


class Client:
    def __init__(self, name, date, address, allergies):
        self.name       = name
        self.address    = address
        self.allergies  = allergies
        self.date       = date


# Functions #

def surplus(dict1, dict2, clientDict):
    # Initializes variables storing the shortest list in the array and its index
    # All three dictionaries must have the exact same keys
    shortest = 10000
    # The loop checks for the shortest list in the array and records its index
    for key in dict1.keys():
        # Determines the total amount of members in a section
        tLen = 0
        for group in dict1[key]:
            tLen += len(group)
        for group in dict2[key]:
            tLen += len(group)
        tLen -= len(clientDict[key])
        # Finds the lowest and returns it to be accessed
        if tLen <= shortest:
            shortest = tLen
            best = key

    return best


def eliminate(volunteer, vDict):
    # This function removes a volunteer from the dictionary
    for key in vDict.keys():
        for group in vDict[key]:
            if group.has_member(volunteer):
                vDict[key].remove(group)
    return vDict


def sFormat(string):
    # Repeats asking for a file until a proper one is given
    if len(string) > 5:
        if string[-5:] == ".xlsx":
            if os.path.isfile(string):
                return string
    return sFormat(input("Error: Please re-enter the name of your file, with the file type(i.e. \"example.xlsx\"):"))
    # Need to edit the phrase

# VARIABLES #
# Dictionaries

# Places the data into groups
firstChoice     = {1: [], 2: [], 3: [], 4: []}
secondChoice    = {1: [], 2: [], 3: [], 4: []}
clients         = {1: [], 2: [], 3: [], 4: []}

# CODE #

print("this function takes excel file and makes another excel file")

wFile = sFormat(input("Enter the name of your volunteer file, with the file type(i.e. \"example.xlsx\"):"))
cFile = sFormat(input("Enter the name of your client file, with the file type(i.e. \"example.xlsx\"):"))

# opens the volunteer workbook
vWorkbook = openpyxl.load_workbook(wFile)
vSheet = vWorkbook["Sheet1"]

# opens the client workbook
cWorkbook = openpyxl.load_workbook(cFile)
cSheet = cWorkbook["Sheet1"]

# Reads the file into the first and second choice dictionaries
for row in range(2, 112):
    names = []
    for letters in range(1, 15, 3):
        j = 65 + letters
        name = vSheet[chr(j)+str(row)].value
        email = vSheet[chr(j+1)+str(row)].value

        if name != None:
            vol = Volunteer(name, email, '', '', [])
            # placeholders, no contact info or phone number present
            names.append(vol)

    # Accesses the first and second choice in the table
    fChoice = int(vSheet[chr(65+15)+str(row)].value)
    sChoice = int(vSheet[chr(65+16)+str(row)].value)

    # Creates the group
    newGroup = Group(names, fChoice, sChoice)

    # Accesses the Q value in the table
    firstChoice[newGroup.day1].append(newGroup)

    # Checks if they make a second choice
    if newGroup.day2 != 5:
        secondChoice[newGroup.day2].append(newGroup)

for row in range(2,22):
    # collects all data from the workbook and creates a client
    name    = cSheet['A'+str(row)].value
    date    = cSheet['G'+str(row)].value
    address = cSheet['I'+str(row)].value
    allergy = cSheet['J'+str(row)].value
    newClient = Client(name, date, address, allergy)
    # appends client to the list on their chosen date
    clients[newClient.date].append(newClient)


masterDict = {}

# main sorting algorithm, repeated 4 times to eliminate all options
for loops in range(4):
    start = surplus(firstChoice, secondChoice, clients)
    vList = firstChoice.pop(start) + secondChoice.pop(start)
    index = 0

    for group in vList:
        destroy = []
        if len(vList)-1 != index:
            if len(group) < 5:
                for add in range(index + 1, len(vList)-1):
                    print(len(vList)-1, add)
                    if len(vList[add]) + len(group) <= 5:
                        vList[index] = group + vList[add]
                        destroy.append(add)
                # deletes the indexes starting from right to left
                for delete in reversed(destroy):
                    del(vList[delete])
        index += 1

    volAssign = 0

    # takes the clients from that day and assigns volunteers
    for client in clients[start]:
        masterDict[client] = vList[volAssign]
        volAssign += 1
        # volunteers that are officially signed are deleted from all other lists
        for volunteer in vList[volAssign].members:
            eliminate(volunteer, firstChoice)
            eliminate(volunteer, secondChoice)

# prints everything into a certain format
for key in masterDict.keys():
    print(key.address, end=":\t\t\t")
    for i in masterDict[key].members:
        print(i.name, end='\t\t')
    print()

# close work books
vWorkbook.close()
cWorkbook.close()
data= ['Saturday AM','Saturday PM','Sunday AM','Sunday PM']
wBook = xlsxwriter.Workbook("schedule.xlsx")
sheet = wBook.add_worksheet()
x = 0
for i in range(0,12,3):
    sheet.write(0,i,data[x])
    sheet.write(0,i+1,'Volunteers')
    sheet.write(0,i+2,'Clients')
    x+=1

row1 =1
row2=1
row3=1
row4=1
for client in clients:
    if client == 1:
        for person in clients[1]:
            namelist = ""
            for item in masterDict:
                for i in range(len(masterDict[person].members)):
                    if (i!=len(masterDict[person].members)-1):
                        namelist += masterDict[person].members[i].name + ", "
                    else:
                        namelist += masterDict[person].members[i].name
            sheet.write(row1,1,namelist)
            sheet.write(row1,2,person.name)
            row1+=1
    if client ==2:

        for person in clients[2]:
            namelist = ""
            for item in masterDict:
                for i in range(len(masterDict[person].members)):
                    if (i != len(masterDict[person].members) - 1):
                        namelist += masterDict[person].members[i].name + ", "
                    else:
                        namelist += masterDict[person].members[i].name
            sheet.write(row2, 4, namelist)
            sheet.write(row2, 5, person.name)
            row2+=1
    if client == 3:
        for person in clients[3]:
            namelist = ""
            for item in masterDict:
                for i in range(len(masterDict[person].members)):
                    if (i != len(masterDict[person].members) - 1):
                        namelist += masterDict[person].members[i].name + ", "
                    else:
                        namelist += masterDict[person].members[i].name
            sheet.write(row3, 7, namelist)
            sheet.write(row3, 8, person.name)
            row3+=1
    if client == 4:
        for person in clients[4]:
            namelist = ""
            for item in masterDict:
                for i in range(len(masterDict[person].members)):
                    if (i != len(masterDict[person].members) - 1):
                        namelist += masterDict[person].members[i].name + ", "
                    else:
                        namelist += masterDict[person].members[i].name
            sheet.write(row4, 10, namelist)
            sheet.write(row4, 11, person.name)
            row4+=1


wBook.close()

