import openpyxl
import os
import xlsxwriter
import smtplib

# Classes #


class Volunteer:
    def __init__(self, name, email, phone, contactMethod, allergies):
        self.name = name
        self.email = email
        self.phone = phone
        self.contactMethod = contactMethod
        self.allergies = allergies


class Group:
    def __init__(self, members, day1, day2):
        self.members = members
        self.day1 = day1
        self.day2 = day2
        self.allergies = []
        # automatically collects allergies from the members and assigns to overall group
        for member in self.members:
            for allergy in member.allergies:
                if allergy not in self.allergies:
                    self.allergies.append(allergy)

    def __len__(self):
        return len(self.members)

    def __add__(self, otherGroup):
        newGroup = Group(self.members + otherGroup.members, self.day1, 5)
        return newGroup

    def has_member(self, volunteer):
        return volunteer in self.members


class Client:
    def __init__(self, name, date, address, allergies,task):
        self.name = name
        self.address = address
        self.allergies = allergies
        self.date = date
        self.task = task

# Functions #

def eliminate(volunteer, vDict):
    # This function removes a volunteer from the dictionary
    for key in vDict.keys():
        for group in vDict[key]:
            if group.has_member(volunteer):
                vDict[key].remove(group)
    return vDict


def help():
    # This function prints basic help and tips
    print("-----------------HELP-----------------\n"
          "1. Ensure that your inputs do not contain unnecessary spaces.\n"
          "2. Inputs are case sensitive. Ensure that you are typing inputs with the correct case.\n"
          "3. Ensure that you type the file name correctly.\n"
          "4. Ensure that the volunteer and client information files are in the specified location.\n"
          "5. Ensure that 'Allow Less Secure Apps' in your Gmail account settings is turned on.\n")
    x = input("--PRESS <Enter> TO CONTINUE--\n\n")

def getClient(dict, name):
    for i in range(1,5):
        for client in dict[i]:
            if client.name == name:
                return client

def getVols(dict, group):
    for key in dict.keys():
        names = []
        for member in dict[key].members:
            names.append(member.name)
        if (group == names):
            return dict[key]



def menu():
    # This code runs the main menu
    print("----------------------------------------------------------------")
    print("|          FIX 'N' CLEAN VOLUNTEER ASSIGNMENT PROGRAM          |")
    print("----------------------------------------------------------------")

    menu_run = True    # Variable Controlling the input check loop

    # This loop runs if the user's input is invalid
    while menu_run == True:

        menu_input = input("Type in the number related to the function you would like to run\n"
                           "1: Volunteer assignment\n2: Automated Email\n3: Help\n4: Quit\n\n->")

        if (menu_input != "1") and (menu_input != "2") and (menu_input != "3") and (menu_input != "4") and (menu_input!="5"):
            print("Invalid input. Please type 1, 2, 3, or 4 depending on which function you would like to access.\n")
        else:
            menu_run = False

    return menu_input


def sFormat(string):
    # Repeats asking for a file until a proper one is given
    if len(string) > 5:
        if string[-5:] == ".xlsx":
            if os.path.isfile(string):
                return string
    return sFormat(input("Error: Please re-enter the name of your file, with the file type(i.e. \"example.xlsx\"):"))
    # Need to edit the phrase


def surplus(dict1, dict2, clientDict):
    # Initializes variables storing the shortest list in the array and its index
    # All three dictionaries must have the exact same keys
    shortest = 10000
    # The loop checks for the shortest list in the array and records its index
    for key in dict1.keys():
        # Determines the total amount of members in a section
        tLen = 0
        for group in dict1[key]:
            tLen += len(group)
        for group in dict2[key]:
            tLen += len(group)
        tLen -= len(clientDict[key])
        # Finds the lowest and returns it to be accessed
        if tLen <= shortest:
            shortest = tLen
            best = key

    return best

# VARIABLES #
# Dictionaries

# Places the data into groups
firstChoice = {1: [], 2: [], 3: [], 4: []}
secondChoice = {1: [], 2: [], 3: [], 4: []}
clients = {1: [], 2: [], 3: [], 4: []}
pRun = True # Controls the main program loop


# CODE #

while (pRun == True):
    # Run the Main Menu
    menu_choice = menu()

    # Based on the user's choice, this logical structure determines which functionality is run
    if (menu_choice == "1"):
        # Runs the Volunteer Assignment function
        print("This function will assign volunteers to time slots.\n")
        wFile = sFormat(input("Enter the name of your volunteer file, with the file type(i.e. \"example.xlsx\"):"))
        cFile = sFormat(input("Enter the name of your client file, with the file type(i.e. \"example.xlsx\"):"))

        # opens the volunteer workbook
        vWorkbook = openpyxl.load_workbook(wFile)
        vSheet = vWorkbook["Sheet1"]

        # opens the client workbook
        cWorkbook = openpyxl.load_workbook(cFile)
        cSheet = cWorkbook["Sheet1"]

        # Reads the file into the first and second choice dictionaries
        for row in range(2, 112):
            names = []
            for letters in range(1, 15, 3):
                j = 65 + letters
                name = vSheet[chr(j) + str(row)].value
                email = vSheet[chr(j + 1) + str(row)].value

                if name != None:
                    vol = Volunteer(name, email, '', '', [])
                    # placeholders, no contact info or phone number present
                    names.append(vol)

            # Accesses the first and second choice in the table
            fChoice = int(vSheet[chr(65 + 15) + str(row)].value)
            sChoice = int(vSheet[chr(65 + 16) + str(row)].value)

            # Creates the group
            newGroup = Group(names, fChoice, sChoice)

            # Accesses the Q value in the table
            firstChoice[newGroup.day1].append(newGroup)

            # Checks if they make a second choice
            if newGroup.day2 != 5:
                secondChoice[newGroup.day2].append(newGroup)

        for row in range(2, 22):
            # collects all data from the workbook and creates a client
            name = cSheet['A' + str(row)].value
            date = cSheet['G' + str(row)].value
            address = cSheet['I' + str(row)].value
            task = cSheet['H' + str(row)].value
            allergy = cSheet['J' + str(row)].value
            newClient = Client(name, date, address, allergy, task)
            # appends client to the list on their chosen date
            clients[newClient.date].append(newClient)

        masterDict = {}

        # main sorting algorithm, repeated 4 times to eliminate all options
        for loops in range(4):
            start = surplus(firstChoice, secondChoice, clients)
            vList = firstChoice.pop(start) + secondChoice.pop(start)
            index = 0

            for group in vList:
                destroy = []
                if len(vList) - 1 != index:
                    if len(group) < 5:
                        for add in range(index + 1, len(vList) - 1):
                            print(len(vList) - 1, add)
                            if len(vList[add]) + len(group) <= 5:
                                vList[index] = group + vList[add]
                                destroy.append(add)
                        # deletes the indexes starting from right to left
                        for delete in reversed(destroy):
                            del (vList[delete])
                index += 1

            volAssign = 0

            # takes the clients from that day and assigns volunteers
            for client in clients[start]:
                masterDict[client] = vList[volAssign]
                volAssign += 1
                # volunteers that are officially signed are deleted from all other lists
                for volunteer in vList[volAssign].members:
                    eliminate(volunteer, firstChoice)
                    eliminate(volunteer, secondChoice)

        # prints everything into a certain format
        for key in masterDict.keys():
            print(key.address, end=":\t\t\t")
            for i in masterDict[key].members:
                print(i.name, end='\t\t')
            print()





        # close work books
        vWorkbook.close()
        cWorkbook.close()
        # headers for 1st sheet
        data = ['Saturday AM', 'Saturday PM', 'Sunday AM', 'Sunday PM']
        # new workbook for writing
        wBook = xlsxwriter.Workbook("schedule.xlsx")
        # sheet with schedule
        sheet = wBook.add_worksheet()
        # sheet with client info
        sheet2 = wBook.add_worksheet()

        # x keeps track of which header to write
        x = 0
        bold = wBook.add_format({'bold': True})

        # adds headers to first row
        for i in range(0, 12, 3):
            sheet.write(0, i, data[x], bold)
            sheet.write(0, i + 1, 'Volunteers', bold)
            sheet.write(0, i + 2, 'Clients', bold)
            x += 1

        # 2D array. 1st array is Sat AM, 2nd is Sat PM, etc.
        cells = [[1, 1, 2], [1, 4, 5], [1, 7, 8], [1, 10, 11]]

        # loop through all days (1-Sat AM, 2-Sat PM, 3-Sun AM, 4-Sun PM)
        for i in range(1, 5):
            # loop through all clients for specific day
            for client in clients[i]:
                # reset list of volunteers
                namelist = ""
                # creates list of names in group separated by comma
                for j in range(len(masterDict[client].members)):
                    if (j != len(masterDict[client].members) - 1):
                        namelist += masterDict[client].members[j].name + ", "
                    else:
                        namelist += masterDict[client].members[j].name
                # write volunteer names and client name to sheet, increment row
                allergyCheck = False
                if len(masterDict[client].allergies) > 0:
                    allergyCheck = True
                abold = bold = wBook.add_format({'bold': allergyCheck})
                sheet.write(cells[i - 1][0], cells[i - 1][1], namelist, abold)
                sheet.write(cells[i - 1][0], cells[i - 1][2], client.name, abold)
                cells[i - 1][0] += 1

        clientHeaders = ['Client Name', 'Address', 'Task', 'Additional Info']
        for i in range(3):
            sheet2.write(0, i, clientHeaders[i], bold)

        row = 1
        for i in range(1, 5):
            for client in clients[i]:
                sheet2.write(row, 0, client.name)
                sheet2.write(row, 1, client.address)
                sheet2.write(row, 2, client.allergies)
                row += 1

        wBook.close()


    elif (menu_choice == '2'):
        wb = openpyxl.load_workbook("schedule.xlsx")
        file = xlsxwriter.Workbook("email_failures.xlsx")
        filesheet = file.add_worksheet()
        fileheaders = ["E-mail Address","Name","Time","Location","Client","Task"]
        for i in range(len(fileheaders)):
            filesheet.write(0,i,fileheaders[i],bold)
        sheet = wb["Sheet1"]
        newDict = {}
        fails = []
        filecol = 1
        vCols = {'B': 0, 'E': 0, 'H': 0, 'K': 0}
        cCols = {'C': 0, 'F': 0, 'I': 0, 'L': 0}
        for col in cCols.keys():
            row = 2
            empty = False
            while not empty:
                if sheet[col + str(row)].value != None:
                    row += 1
                    cCols[col] += 1
                else:
                    empty = True

        for col in cCols.keys():
            x = 1
            row = 2
            for i in range(cCols[col]):
                clientName = sheet[col + str(row)].value
                client = getClient(clients, clientName)
                vols = sheet[chr(ord(col) - 1) + str(row)].value.split(", ")
                group = getVols(masterDict, vols)
                newDict[client] = group
                row += 1

        wb.close()
        sender = input("\nPlease enter your Gmail email address\n->")
        username = sender
        password = input("\nPlease enter the password to your Gmail account (this information will not be permanently saved/stored)\n->")

        for i in range(1, 5):
            for client in clients[i]:
                for j in range(len(newDict[client].members)):
                    name = newDict[client].members[j].name.split(" ")[0]

                    # receivers = masterDict[client].members[j].email
                    receivers = masterDict[client].members[j].email
                    address = client.address
                    task = client.task
                    message = """From: Fix 'N' Clean Team %s
                           To: %s
                           Subject: Fix 'N' Clean Volunteer Assignment

                           Hi %s,
                           This is a message notifying you of your Fix 'N' Clean Volunteer Assignment.
                           You will be volunteering on %s at %s for %s.
                           Here is what your client wants you to do\n:
                           %s
                           Have a good day,
                           Fix 'N' Clean Management Team
                           """ % (
                    sender, newDict[client].members[j].email, newDict[client].members[j].name, data[i - 1],
                    client.address, client.name, client.task)

                    # AUTHENTICATE
                    try:
                        smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
                        smtpObj.ehlo()
                        smtpObj.starttls()
                        smtpObj.login(username, password)
                        smtpObj.sendmail(sender, receivers, message)
                        # print("Successfully sent email")
                        x += 1
                        smtpObj.close()
                    except:
                        print("Error: unable to send email")
                        missed = [sender, newDict[client].members[j].email, newDict[client].members[j].name, data[i - 1],
                    client.address, client.name, client.task]
                        fails.append(missed)
                        for k in range(len(missed)):
                            filesheet.write(filecol,k,missed[k])
                        filecol += 1



        print(x)

    elif (menu_choice == "3"):
        # Runs the Help function
        help()

    elif (menu_choice == "4"):
        # Ends the progrm loop to close the program
        pRun = False





# Closes the program once the main program loop ends

exit(0)